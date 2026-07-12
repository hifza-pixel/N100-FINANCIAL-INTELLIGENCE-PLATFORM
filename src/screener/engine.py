from pathlib import Path
import sqlite3
import os
import pandas as pd
import yaml
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parents[2]

DATABASE = BASE_DIR / "db" / "nifty100.db"

CONFIG = BASE_DIR / "config" / "screener_config.yaml"

def load_config():
    """
    Load YAML Configuration
    """

    with open(CONFIG, "r") as file:
        return yaml.safe_load(file)


def get_connection():
    return sqlite3.connect(DATABASE)


def load_financial_ratios():

    print("Database Path:", DATABASE)
    print("Database Exists:", DATABASE.exists())

    conn = get_connection()

    print(pd.read_sql("SELECT COUNT(*) AS total FROM financial_ratios", conn))

    df = pd.read_sql(
        "SELECT * FROM financial_ratios",
        conn,
    )

    conn.close()

    print("Rows Loaded:", len(df))

    return df


# ==============================
# Helper
# ==============================

def has_column(df, column):
    """
    Check whether column exists.
    """

    return column in df.columns


# ==============================
# Generic Filter Engine
# ==============================

def apply_filters(df, filters):
    """
    Apply filters dynamically.
    """

    result = df.copy()

    # --------------------------
    # ROE
    # --------------------------

    if (
        "roe_min" in filters
        and has_column(result, "return_on_equity_pct")
    ):

        result = result[
            result["return_on_equity_pct"]
            >= filters["roe_min"]
        ]

    # --------------------------
    # Debt to Equity
    # --------------------------

    if (
        "debt_to_equity_max" in filters
        and has_column(result, "debt_to_equity")
    ):

        result = result[
            result["debt_to_equity"]
            <= filters["debt_to_equity_max"]
        ]

    # --------------------------
    # Free Cash Flow
    # --------------------------

    if (
        "free_cash_flow_min" in filters
        and has_column(result, "free_cash_flow_cr")
    ):

        result = result[
            result["free_cash_flow_cr"]
            >= filters["free_cash_flow_min"]
        ]

    # --------------------------
    # Operating Profit Margin
    # --------------------------

    if (
        "operating_profit_margin_min" in filters
        and has_column(result, "operating_profit_margin_pct")
    ):

        result = result[
            result["operating_profit_margin_pct"]
            >= filters["operating_profit_margin_min"]
        ]

    # --------------------------
    # Interest Coverage
    # --------------------------

    if (
        "interest_coverage_min" in filters
        and has_column(result, "interest_coverage")
    ):

        result = result[
            (
                result["interest_coverage"]
                >= filters["interest_coverage_min"]
            )
            |
            (
                result["interest_coverage"].isna()
            )
        ]

    # --------------------------
    # Asset Turnover
    # --------------------------

    if (
        "asset_turnover_min" in filters
        and has_column(result, "asset_turnover")
    ):

        result = result[
            result["asset_turnover"]
            >= filters["asset_turnover_min"]
        ]

    # --------------------------
    # Sales
    # --------------------------

    if (
        "sales_min" in filters
        and has_column(result, "sales")
    ):

        result = result[
            result["sales"]
            >= filters["sales_min"]
        ]

    # --------------------------
    # Net Profit
    # --------------------------

    if (
        "net_profit_min" in filters
        and has_column(result, "net_profit")
    ):

        result = result[
            result["net_profit"]
            >= filters["net_profit_min"]
        ]
    return result.reset_index(drop=True)

def normalize(series):
    """
    Normalize values between 0 and 100.
    """

    if series.empty:
        return series

    minimum = series.min()
    maximum = series.max()

    if minimum == maximum:
        return pd.Series(
            [50] * len(series),
            index=series.index,
        )

    return (
        (
            (series - minimum)
            /
            (maximum - minimum)
        ) * 100
    )

def winsorize(series):
    """
    Clip values between 10th and 90th percentile.
    """

    p10 = series.quantile(0.10)
    p90 = series.quantile(0.90)

    return series.clip(lower=p10, upper=p90)
def calculate_composite_score(df):
   
    result = df.copy()

    required = [

        "return_on_equity_pct",

        "return_on_capital_employed_pct",

        "net_profit_margin_pct",

        "free_cash_flow_cr",

        "debt_to_equity",

        "interest_coverage",

        "asset_turnover",

    ]

    # -----------------------
    # Missing Columns
    # -----------------------

    for col in required:

        if col not in result.columns:

            result[col] = 0

    # -----------------------
    # Convert Numeric
    # -----------------------

    for col in required:

        result[col] = pd.to_numeric(
            result[col],
            errors="coerce",
        ).fillna(0)

    # -----------------------
    # Winsorisation
    # -----------------------

    for col in required:

        result[col] = winsorize(
            result[col]
        )

    # -----------------------
    # Normalize Scores
    # -----------------------

    result["roe_score"] = normalize(
        result["return_on_equity_pct"]
    )

    result["roce_score"] = normalize(
        result["return_on_capital_employed_pct"]
    )

    result["npm_score"] = normalize(
        result["net_profit_margin_pct"]
    )

    result["fcf_score"] = normalize(
        result["free_cash_flow_cr"]
    )

    result["asset_score"] = normalize(
        result["asset_turnover"]
    )

    result["icr_score"] = normalize(
        result["interest_coverage"]
    )

    # -----------------------
    # Debt Score
    # Lower is Better
    # -----------------------

    result["de_score"] = 100 - normalize(
        result["debt_to_equity"]
    )

    # -----------------------
    # Growth Scores
    # -----------------------

    if "revenue_cagr_5yr" in result.columns:

        result["revenue_growth_score"] = normalize(

            winsorize(

                pd.to_numeric(
                    result["revenue_cagr_5yr"],
                    errors="coerce",
                ).fillna(0)

            )

        )

    else:

        result["revenue_growth_score"] = 50

    if "pat_cagr_5yr" in result.columns:

        result["pat_growth_score"] = normalize(

            winsorize(

                pd.to_numeric(
                    result["pat_cagr_5yr"],
                    errors="coerce",
                ).fillna(0)

            )

        )

    else:

        result["pat_growth_score"] = 50

    # -----------------------
    # Final Weighted Score
    # -----------------------

    result["composite_quality_score"] = (

        result["roe_score"] * 0.15

        + result["roce_score"] * 0.10

        + result["npm_score"] * 0.10

        + result["fcf_score"] * 0.30

        + result["revenue_growth_score"] * 0.10

        + result["pat_growth_score"] * 0.10

        + result["de_score"] * 0.10

        + result["icr_score"] * 0.03

        + result["asset_score"] * 0.02

    )

    # -----------------------
    # Final Normalize
    # -----------------------

    result["composite_quality_score"] = normalize(
        result["composite_quality_score"]
    ).round(2)

    return result
def apply_preset(df, config, preset_name):
   

    presets = config["filters"]

    if preset_name not in presets:

        raise ValueError(
            f"{preset_name} not found."
        )

    preset = presets[preset_name]

    result = apply_filters(
        df,
        preset,
    )

    result = calculate_composite_score(result)

    result = result.sort_values(
        by="composite_quality_score",
        ascending=False,
    ).reset_index(drop=True)

    return result


# ==============================
# Display Helper
# ==============================

def show_result(title, df):

    print("\n" + "=" * 60)

    print(title)

    print("=" * 60)

    print(f"Companies : {len(df)}")

    if len(df) == 0:

        print("No companies found.")

        return

    columns = [

        "company_id",

        "year",

        "return_on_equity_pct",

        "debt_to_equity",

        "composite_quality_score",

    ]

    available = [

        c

        for c in columns

        if c in df.columns

    ]

    print(df[available].head(10))
    # ==============================
# Main
# ==============================

def main():

    print("=" * 70)
    print("Sprint 3 - Screener Engine")
    print("=" * 70)

    # -------------------------
    # Load Config
    # -------------------------

    config = load_config()


    df = load_financial_ratios()

    print("\nOriginal Rows :", len(df))
    if "default" in config["filters"]:

        filtered = apply_filters(
            df,
            config["filters"]["default"],
        )

        filtered = calculate_composite_score(filtered)

        filtered = filtered.sort_values(
            by="composite_quality_score",
            ascending=False,
        )

        show_result(
            "Default Screener",
            filtered,
        )
    presets = [
        "quality_compounder",
        "value_pick",
        "growth_accelerator",
        "dividend_champion",
        "debt_free_blue_chip",
        "turnaround_watch",

    ]
    results = {}
    for preset in presets:

        try:

            result = apply_preset(df,config,preset, )
            results[preset] = result
            show_result(
                preset.replace("_", " ").title(),
                result,
            )
        except Exception as e:
            print(f"\n {preset}")
            print(e)

    print("\n" + "=" * 70)
    print("All Screeners Completed")
    print("=" * 70)
    export_screeners(results)
    print("\nSprint 3 Day 17 Completed")
    return results

OUTPUT_DIR = BASE_DIR / "output"


def export_screeners(results):
    """
    Export all preset screeners to Excel
    Sprint 3 – Day 17
    """

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    output_file = OUTPUT_DIR / "screener_output.xlsx"

    # ===============================
    # Excel Writer
    # ===============================

    with pd.ExcelWriter(
        output_file,
        engine="openpyxl",
    ) as writer:

        # ===============================
        # Presets
        # ===============================

        for preset, df in results.items():

            if df.empty:
                continue

            export_columns = [

                "company_id",
                "year",

                "return_on_equity_pct",
                "return_on_capital_employed_pct",

                "net_profit_margin_pct",

                "debt_to_equity",

                "interest_coverage",

                "asset_turnover",

                "free_cash_flow_cr",

                "earnings_per_share",

                "book_value_per_share",

                "dividend_payout_ratio_pct",

                "cash_from_operations_cr",

                "composite_quality_score",

            ]

            export_columns = [

                col

                for col in export_columns

                if col in df.columns

            ]

            export_df = df[
                export_columns
            ]

            sheet = preset[:31]

            export_df.to_excel(

                writer,

                sheet_name=sheet,

                index=False,

            )

            worksheet = writer.sheets[sheet]

            # ====================================
            # Header Formatting
            # ====================================

            header_fill = PatternFill(

                fill_type="solid",

                start_color="1F4E78",

                end_color="1F4E78",

            )

            header_font = Font(

                bold=True,

                color="FFFFFF",

            )

            for cell in worksheet[1]:

                cell.fill = header_fill

                cell.font = header_font

            # ====================================
            # Freeze Header
            # ====================================

            worksheet.freeze_panes = "A2"

            # ====================================
            # Auto Width
            # ====================================

            for column in worksheet.columns:

                max_length = 0

                column_letter = get_column_letter(

                    column[0].column

                )

                for cell in column:

                    try:

                        if len(str(cell.value)) > max_length:

                            max_length = len(

                                str(cell.value)

                            )

                    except:

                        pass

                worksheet.column_dimensions[
                    column_letter
                ].width = max_length + 3

            # ====================================
            # Green / Red Formatting
            # ====================================

            green = PatternFill(

                fill_type="solid",

                start_color="C6EFCE",

                end_color="C6EFCE",

            )

            red = PatternFill(

                fill_type="solid",

                start_color="FFC7CE",

                end_color="FFC7CE",

            )

            headers = {}

            for cell in worksheet[1]:

                headers[cell.value] = cell.column

            for row in range(

                2,

                worksheet.max_row + 1,

            ):

                # ----------------
                # ROE
                # ----------------

                if "return_on_equity_pct" in headers:

                    c = worksheet.cell(

                        row,

                        headers[
                            "return_on_equity_pct"
                        ],

                    )

                    if c.value is not None:

                        if c.value >= 15:

                            c.fill = green

                        else:

                            c.fill = red

                # ----------------
                # Debt
                # ----------------

                if "debt_to_equity" in headers:

                    c = worksheet.cell(

                        row,

                        headers[
                            "debt_to_equity"
                        ],

                    )

                    if c.value is not None:

                        if c.value <= 1:

                            c.fill = green

                        else:

                            c.fill = red

                # ----------------
                # Free Cash Flow
                # ----------------

                if "free_cash_flow_cr" in headers:

                    c = worksheet.cell(

                        row,

                        headers[
                            "free_cash_flow_cr"
                        ],

                    )

                    if c.value is not None:

                        if c.value > 0:

                            c.fill = green

                        else:

                            c.fill = red

                # ----------------
                # Composite Score
                # ----------------

                if "composite_quality_score" in headers:

                    c = worksheet.cell(

                        row,

                        headers[
                            "composite_quality_score"
                        ],

                    )

                    if c.value is not None:

                        if c.value >= 70:

                            c.fill = green

                        elif c.value < 40:

                            c.fill = red

    print("\n" + "=" * 60)

    print("Excel Export Completed")

    print("=" * 60)

    print(f"File Saved : {output_file}")

    return output_file


if __name__ == "__main__":
    results = main()